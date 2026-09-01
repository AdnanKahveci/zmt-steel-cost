from __future__ import annotations

import hashlib
import json
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = next(ROOT.glob("*.xlsx"))
LEGACY_DIR = ROOT / "Legacy"
DOCS_DIR = ROOT / "Docs"
LEGACY_COPY = LEGACY_DIR / SOURCE.name
BASELINE_PATH = LEGACY_DIR / "LegacyBaseline.json"
FORMULA_JSON_PATH = LEGACY_DIR / "FormulaCatalog.json"
MATERIAL_JSON_PATH = LEGACY_DIR / "MaterialCatalog.json"

CATEGORY_NAMES = {
    1001: "Hafif Çelik Panel ve Metal Aksam",
    1002: "Alçıpan ve Kaplama",
    1003: "Çatı Sacı",
    1004: "Kapı ve Pencere",
    1005: "Elektrik",
    1006: "Vida",
    1007: "Depo ve Hırdavat",
    1008: "Sıhhi Tesisat",
    1009: "Çatı Oluğu ve Boru",
    1010: "Boya ve Mastik",
}

EXPECTED_GROUP_TOTALS = {
    1001: Decimal("342767.9849784"),
    1002: Decimal("298366.872"),
    1003: Decimal("78615.26587834278"),
    1004: Decimal("165634.9056"),
    1005: Decimal("34845.8157"),
    1006: Decimal("15185.94"),
    1007: Decimal("75190.950650112"),
    1008: Decimal("52706.23536"),
    1009: Decimal("7483.36744"),
    1010: Decimal("42678.408"),
}


@dataclass(frozen=True)
class CategoryRange:
    code: int
    header_row: int
    first_material_row: int
    last_material_row: int
    total_row: int


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def invariant(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        return format(value, ".17g")
    if isinstance(value, Decimal):
        return format(value, "f")
    return value


def markdown_value(value: Any) -> str:
    if value is None:
        return "—"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).replace("\r", " ").replace("\n", " ⏎ ").replace("|", "\\|")


def code_name(value: str) -> str:
    replacements = str.maketrans(
        "çğıöşüÇĞİÖŞÜ",
        "cgiosuCGIOSU",
    )
    words = re.findall(r"[A-Za-z0-9]+", value.translate(replacements))
    return "".join(word[:1].upper() + word[1:].lower() for word in words) or "Rule"


def dependency_list(formula: str) -> list[str]:
    # Captures local and qualified A1 references. Ranges remain readable as two endpoints.
    pattern = re.compile(
        r"(?:(?:'[^']+'|[A-Za-zÇĞİÖŞÜçğıöşü ]+)!){0,1}\$?[A-Z]{1,3}\$?\d+"
    )
    seen: list[str] = []
    for item in pattern.findall(formula or ""):
        normalized = item.replace("$", "").strip()
        if normalized not in seen:
            seen.append(normalized)
    return seen


def formula_rule(sheet_name: str, coordinate: str, material_code: str | None) -> str:
    if material_code:
        column = re.match(r"[A-Z]+", coordinate).group(0)
        member = {
            "D": "CalculateQuantity",
            "E": "CalculateSalesUnitPrice",
            "F": "CalculateSalesLineTotal",
            "I": "CalculatePurchaseUnitPriceExVat",
            "J": "CalculatePurchaseLineTotalExVat",
            "K": "CalculatePurchaseUnitPriceIncVat",
            "L": "CalculateDiscountedSalesUnitPrice",
            "M": "CalculateGrossMarginRate",
        }.get(column, "CalculateMaterialValue")
        return f"LegacyExcelV1Rules.{member}(\"{material_code}\", context)"
    if sheet_name == "İSİMLENDİRME" and coordinate in {"Q17", "R17", "S17"}:
        return "RoofCalculationService (slope factor / eave area / roof cover area)"
    if sheet_name == "BİNA BİLGİLERİ" and coordinate.startswith("F"):
        return "ScopeTotalsService"
    return f"LegacyExcelV1WorkbookRules.{code_name(sheet_name)}{coordinate}"


def parse_validations(source: Path, sheet_names: list[str]) -> list[dict[str, str]]:
    results: list[dict[str, str]] = []
    with ZipFile(source) as archive:
        for index, sheet_name in enumerate(sheet_names, start=1):
            xml_name = f"xl/worksheets/sheet{index}.xml"
            if xml_name not in archive.namelist():
                continue
            root = etree.fromstring(archive.read(xml_name))
            for validation in root.xpath('//*[local-name()="dataValidation"]'):
                formulas = validation.xpath('.//*[local-name()="formula1"]/*[local-name()="f"]/text()')
                if not formulas:
                    formulas = validation.xpath('./*[local-name()="formula1"]/text()')
                sqrefs = validation.xpath('.//*[local-name()="sqref"]/text()')
                target = sqrefs[0] if sqrefs else validation.get("sqref", "")
                results.append(
                    {
                        "sheet": sheet_name,
                        "target": target,
                        "type": validation.get("type", ""),
                        "formula": formulas[0] if formulas else "",
                    }
                )
    return results


def category_ranges(sheet: Any) -> list[CategoryRange]:
    ranges: list[CategoryRange] = []
    current_code: int | None = None
    header = 0
    material_rows: list[int] = []
    for row in range(1, sheet.max_row + 1):
        column_d = sheet.cell(row, 4).value
        if isinstance(column_d, int) and column_d in CATEGORY_NAMES:
            current_code = column_d
            header = row
            material_rows = []
            continue
        if current_code is None:
            continue
        if isinstance(sheet.cell(row, 1).value, (int, float)) and sheet.cell(row, 2).value:
            material_rows.append(row)
        if str(column_d).strip().upper() == "TOPLAM":
            ranges.append(
                CategoryRange(
                    current_code,
                    header,
                    material_rows[0],
                    material_rows[-1],
                    row,
                )
            )
            current_code = None
    return ranges


def material_code(category: int, ordinal: int) -> str:
    return f"{category}-{ordinal:03d}"


def extract() -> None:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    LEGACY_DIR.mkdir(parents=True, exist_ok=True)

    source_hash = sha256(SOURCE)
    if not LEGACY_COPY.exists() or sha256(LEGACY_COPY) != source_hash:
        shutil.copy2(SOURCE, LEGACY_COPY)
    if sha256(LEGACY_COPY) != source_hash:
        raise RuntimeError("Legacy copy hash mismatch")

    formula_book = load_workbook(SOURCE, data_only=False, read_only=False)
    value_book = load_workbook(SOURCE, data_only=True, read_only=False)
    offer = formula_book.worksheets[1]
    offer_values = value_book.worksheets[1]
    categories = category_ranges(offer)
    validations = parse_validations(SOURCE, formula_book.sheetnames)

    material_by_row: dict[int, dict[str, Any]] = {}
    materials: list[dict[str, Any]] = []
    for category in categories:
        ordinal = 0
        for row in range(category.first_material_row, category.last_material_row + 1):
            if not (isinstance(offer.cell(row, 1).value, (int, float)) and offer.cell(row, 2).value):
                continue
            ordinal += 1
            code = material_code(category.code, ordinal)
            quantity_formula = offer.cell(row, 4).value if offer.cell(row, 4).data_type == "f" else None
            purchase_formula = offer.cell(row, 9).value if offer.cell(row, 9).data_type == "f" else None
            sales_formula = offer.cell(row, 5).value if offer.cell(row, 5).data_type == "f" else None
            item = {
                "id": len(materials) + 1,
                "code": code,
                "categoryId": category.code,
                "category": CATEGORY_NAMES[category.code],
                "sortOrder": ordinal,
                "excelRow": row,
                "legacyNumber": invariant(offer.cell(row, 1).value),
                "name": offer.cell(row, 2).value.strip() if isinstance(offer.cell(row, 2).value, str) else offer.cell(row, 2).value,
                "specification": invariant(offer.cell(row, 3).value),
                "unit": offer.cell(row, 3).value if str(offer.cell(row, 3).value or "").upper() in {"KG", "ADET", "M2", "M²", "MT"} else "adet",
                "quantityRuleId": f"LegacyExcel-v1:{code}:Quantity",
                "pricingRuleId": f"LegacyExcel-v1:{code}:Pricing",
                "isActive": True,
                "allowManualQuantityOverride": True,
                "allowManualPriceOverride": True,
                "quantityFormula": quantity_formula,
                "quantityConstant": None if quantity_formula else invariant(offer.cell(row, 4).value),
                "purchasePriceFormula": purchase_formula,
                "purchasePriceConstant": None if purchase_formula else invariant(offer.cell(row, 9).value),
                "salesPriceFormula": sales_formula,
                "salesPriceConstant": None if sales_formula else invariant(offer.cell(row, 5).value),
                "expectedQuantity": invariant(offer_values.cell(row, 4).value or 0),
                "expectedSalesUnitPrice": invariant(offer_values.cell(row, 5).value or 0),
                "expectedSalesLineTotal": invariant(offer_values.cell(row, 6).value or 0),
                "expectedPurchaseUnitPriceExVat": invariant(offer_values.cell(row, 9).value or 0),
                "expectedPurchaseLineTotalExVat": invariant(offer_values.cell(row, 10).value or 0),
                "expectedPurchaseUnitPriceIncVat": invariant(offer_values.cell(row, 11).value or 0),
                "expectedDiscountedSalesUnitPrice": invariant(offer_values.cell(row, 12).value or 0),
            }
            materials.append(item)
            material_by_row[row] = item

    if len(materials) != 186:
        raise RuntimeError(f"Expected 186 materials, found {len(materials)}")

    formula_entries: list[dict[str, Any]] = []
    for sheet in formula_book.worksheets:
        values = value_book[sheet.title]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue
                material = material_by_row.get(cell.row) if sheet is offer else None
                formula_entries.append(
                    {
                        "sheet": sheet.title,
                        "cell": cell.coordinate,
                        "materialCode": material["code"] if material else None,
                        "material": material["name"] if material else None,
                        "category": material["category"] if material else None,
                        "formula": cell.value,
                        "dependencies": dependency_list(cell.value),
                        "expected": invariant(values[cell.coordinate].value),
                        "csharpRule": formula_rule(sheet.title, cell.coordinate, material["code"] if material else None),
                    }
                )

    group_totals: list[dict[str, Any]] = []
    for category in categories:
        actual = Decimal(str(offer_values.cell(category.total_row, 6).value or 0))
        expected = EXPECTED_GROUP_TOTALS[category.code]
        group_totals.append(
            {
                "categoryId": category.code,
                "category": CATEGORY_NAMES[category.code],
                "totalCell": f"TEKLİF!F{category.total_row}",
                "expectedSalesTotal": invariant(actual),
                "specifiedRegressionValue": invariant(expected),
                "difference": invariant(abs(actual - expected)),
            }
        )

    inputs_sheet = formula_book.worksheets[0]
    input_values_sheet = value_book.worksheets[0]
    input_cells = [
        "B2", "B3", "B4", "B5", "B6", "B7", "B8", "B9", "B10", "B11", "B12", "B13", "B14",
        "B16", "B17", "B18", "B21", "B22", "B23", "B24", "B25", "B26", "B29", "B30",
        "B32", "D32", "E32", "F32", "B33", "D33", "E33", "F33", "B34", "D34", "E34", "F34",
        "B35", "D35", "E35", "F35", "B36",
        "B39", "B40", "B41", "B42", "B43",
        "B46", "B47", "B48", "B49", "B50", "B51", "B52", "B53", "B54", "B55", "B56", "B57", "C46",
        "B61", "E61", "B62", "E62", "B63", "E63", "B64", "E64",
    ]
    inputs = {
        coordinate: invariant(input_values_sheet[coordinate].value)
        for coordinate in input_cells
    }

    baseline = {
        "schemaVersion": 1,
        "formulaVersion": "LegacyExcel-v1",
        "sourceWorkbook": SOURCE.name,
        "sourceSha256": source_hash,
        "input": inputs,
        "pricingParameters": {
            "exchangeRate": invariant(offer_values["G2"].value),
            "steelPrice": invariant(offer_values["H2"].value),
            "sSeriesPrice": invariant(offer_values["N2"].value),
            "galvanizedPrice": invariant(offer_values["O2"].value),
            "paintedSheetPrice": invariant(offer_values["P2"].value),
            "salesMarkupFactor": "1.73",
            "purchaseVatRate": "0.20",
            "buildingSummaryDiscountRate": "0.21",
            "offerDiscountRate": "0.25",
            "salesVatRate": "0",
        },
        "materials": materials,
        "categoryTotals": group_totals,
        "totals": {
            "fullCalculatedValue": invariant(offer_values["F222"].value),
            "offerDiscountAmount": invariant(offer_values["F223"].value),
            "offerAfterDiscount": invariant(offer_values["F224"].value),
            "offerVat": invariant(offer_values["F225"].value),
            "offerGrandTotal": invariant(offer_values["F226"].value),
            "supplierScopeValue": invariant(input_values_sheet["F86"].value),
            "buildingSummaryDiscountAmount": invariant(input_values_sheet["F87"].value),
            "buildingSummaryAfterDiscount": invariant(input_values_sheet["F88"].value),
            "buildingSummaryVat": invariant(input_values_sheet["F89"].value),
            "buildingSummaryGrandTotal": invariant(input_values_sheet["C90"].value),
        },
    }

    BASELINE_PATH.write_text(json.dumps(baseline, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    MATERIAL_JSON_PATH.write_text(json.dumps(materials, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    FORMULA_JSON_PATH.write_text(json.dumps(formula_entries, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    write_excel_analysis(formula_book, value_book, source_hash, categories, validations, materials, formula_entries)
    write_formula_catalog(formula_entries)
    write_material_catalog(categories, materials)
    write_lookup_catalog(validations, formula_book)

    print(f"Workbook SHA-256: {source_hash}")
    print(f"Materials: {len(materials)}")
    print(f"Formula cells: {len(formula_entries)}")
    print(f"Validations: {len(validations)}")
    print(f"Legacy copy: {LEGACY_COPY}")


def write_excel_analysis(
    formula_book: Any,
    value_book: Any,
    source_hash: str,
    categories: list[CategoryRange],
    validations: list[dict[str, str]],
    materials: list[dict[str, Any]],
    formula_entries: list[dict[str, Any]],
) -> None:
    lines = [
        "# Excel Analizi",
        "",
        "> Bu belge `tools/extract_legacy_workbook.py` ile kaynak çalışma kitabından deterministik olarak üretilmiştir.",
        "",
        "## Kaynak ve değişmezlik",
        "",
        f"- Kaynak: `{SOURCE.name}`",
        f"- Legacy kopya: `Legacy/{SOURCE.name}`",
        f"- SHA-256: `{source_hash}`",
        "- Çalışma zamanında Excel/Office/Interop kullanılmayacaktır.",
        "- Cached değerler Golden Master olarak `Legacy/LegacyBaseline.json` içine alınmıştır.",
        "",
        "## Sayfa envanteri",
        "",
        "| Sayfa | Durum | Kullanılan aralık | Dolu hücre | Formül | Birleşik aralık | Doğrulama |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    for sheet in formula_book.worksheets:
        nonempty = sum(1 for row in sheet.iter_rows() for cell in row if cell.value is not None)
        formulas = sum(1 for row in sheet.iter_rows() for cell in row if cell.data_type == "f")
        validation_count = sum(1 for item in validations if item["sheet"] == sheet.title)
        lines.append(
            f"| {sheet.title} | {sheet.sheet_state} | {sheet.calculate_dimension()} | {nonempty} | {formulas} | {len(sheet.merged_cells.ranges)} | {validation_count} |"
        )

    lines += [
        "",
        "## Ana girdiler",
        "",
        "`BİNA BİLGİLERİ` sayfasındaki sabit/cached girdiler aşağıdadır. Boş değerler legacy örnekte gerçekten boştur.",
        "",
        "| Hücre | Etiket / anlam | Değer | Birim |",
        "|---|---|---:|---|",
    ]
    input_sheet = formula_book.worksheets[0]
    input_values = value_book.worksheets[0]
    units = {
        "B2": "m²", "B3": "kg/m²", "B4": "adet", "B5": "m", "B6": "m", "B7": "kat",
        "B8": "m²", "B9": "m", "B11": "%", "B12": "m²", "B16": "m", "B17": "m", "B18": "m",
        "B29": "m", "B30": "m²", "B32": "m", "B33": "m", "B34": "m²", "B35": "m²", "B36": "m²",
    }
    for row in range(1, input_sheet.max_row + 1):
        label = input_sheet.cell(row, 1).value
        if label is None:
            continue
        for column in range(2, 7):
            cell = input_sheet.cell(row, column)
            if cell.value is not None and (column == 2 or row in {3, 31, 32, 33, 34, 35, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 60, 61, 62, 63, 64}):
                coordinate = cell.coordinate
                lines.append(
                    f"| {coordinate} | {markdown_value(label)} | {markdown_value(input_values[coordinate].value)} | {units.get(coordinate, markdown_value(input_sheet.cell(row, 3).value) if column == 2 else '—')} |"
                )

    lines += [
        "",
        "## Malzeme grupları",
        "",
        "| Kod | Grup | Header | Malzeme satırları | Toplam hücresi | Satır sayısı |",
        "|---:|---|---:|---:|---:|---:|",
    ]
    for category in categories:
        count = sum(1 for item in materials if item["categoryId"] == category.code)
        lines.append(
            f"| {category.code} | {CATEGORY_NAMES[category.code]} | TEKLİF!{category.header_row} | {category.first_material_row}–{category.last_material_row} | TEKLİF!F{category.total_row} | {count} |"
        )
    lines += [
        "",
        f"Toplam **{len(materials)}** malzeme satırı bulunmuştur; hiçbir satır atlanmamıştır.",
        "",
        "## Dropdown / liste doğrulamaları",
        "",
        "| Sayfa | Hedef | Tür | Kaynak / değerler |",
        "|---|---|---|---|",
    ]
    for item in validations:
        lines.append(f"| {item['sheet']} | {item['target']} | {item['type']} | {markdown_value(item['formula'])} |")

    lines += [
        "",
        "## Fiyat ve katsayı kaynakları",
        "",
        "| Parametre | Hücre | Cached değer | Uygulama modeli |",
        "|---|---|---:|---|",
        f"| USD/TL | TEKLİF!G2 | {markdown_value(value_book.worksheets[1]['G2'].value)} | `PricingParameters.ExchangeRate` |",
        f"| Çelik fiyatı | TEKLİF!H2 | {markdown_value(value_book.worksheets[1]['H2'].value)} | `PricingParameters.SteelPrice` |",
        f"| S seri | TEKLİF!N2 | {markdown_value(value_book.worksheets[1]['N2'].value)} | `PricingParameters.SSeriesPrice` |",
        f"| Galvaniz | TEKLİF!O2 | {markdown_value(value_book.worksheets[1]['O2'].value)} | `PricingParameters.GalvanizedPrice` |",
        f"| Boyalı sac | TEKLİF!P2 | {markdown_value(value_book.worksheets[1]['P2'].value)} | `PricingParameters.PaintedSheetPrice` |",
        "| Satış katsayısı | TEKLİF formülleri | 1.73 | `PricingParameters.SalesMarkupFactor` |",
        "| Alış KDV | TEKLİF formülleri | 20% | `PricingParameters.PurchaseVatRate` |",
        "| Bina özeti iskonto | BİNA BİLGİLERİ!B87 | 21% | Legacy parity alanı |",
        "| Teklif iskonto | TEKLİF!E223 | 25% | Legacy parity alanı |",
        "| Satış KDV | BİNA BİLGİLERİ!B89 / TEKLİF!E225 | boş → 0% | Proje override alanı |",
        "",
        "## Çatı katsayıları",
        "",
        "| Eğim | Katsayı |",
        "|---:|---:|",
    ]
    lookup = value_book.worksheets[2]
    for row in range(19, 27):
        lines.append(f"| {Decimal(str(lookup.cell(row,16).value))*100}% | {lookup.cell(row,17).value} |")

    lines += [
        "",
        "## Formül bağımlılıkları",
        "",
        f"Toplam {len(formula_entries)} formül hücresi vardır. Hücre bazlı formül, bağımlılık, cached sonuç ve C# hedef kural eşlemesi `Docs/FormulaCatalog.md` ve makine okunur `Legacy/FormulaCatalog.json` dosyalarındadır.",
        "",
        "Ana akış:",
        "",
        "```text",
        "BİNA BİLGİLERİ girdileri",
        "  ├─ İSİMLENDİRME!P17:S17 → çatı eğim katsayısı ve kaplama alanı",
        "  ├─ TEKLİF!D10:D220 → 186 malzeme miktarı",
        "  ├─ TEKLİF!I/K/E → alış, KDV dahil alış ve satış birim fiyatı",
        "  ├─ TEKLİF!F → satır satış toplamları",
        "  ├─ TEKLİF!F37…F221 → 10 grup toplamı",
        "  ├─ TEKLİF!F222:F226 → tüm gruplar/iskonto/KDV/genel toplam",
        "  └─ BİNA BİLGİLERİ!F76:F90 → sorumluluk kapsamı/iskonto/KDV/genel toplam",
        "```",
        "",
        "## Legacy / Unused",
        "",
        "`FORMÜL` sayfası hidden durumdadır. Aktif sayfalardaki formüllerin hiçbirinde `FORMÜL!` referansı yoktur; bu nedenle ana hesap motoruna dahil edilmemiştir. Sayfadaki 52 formül kaybolmaması için kataloglanmış, ancak `Legacy / Unused` olarak sınıflandırılmıştır.",
        "",
        "## Bilinen legacy tutarsızlıkları",
        "",
        "- BİNA BİLGİLERİ özeti yalnız ZMT kapsamını; TEKLİF sayfası bütün grupları toplar. Uygulamada ayrı değerlerdir.",
        "- Bina özeti %21, TEKLİF %25 iskonto kullanır. Parity testleri ikisini ayrı doğrular.",
        "- Satış KDV girişi boştur ve cached sonuç 0'dır; yeni UI bunu açık bir proje alanı yapar.",
        "- Bazı legacy kâr hücrelerinde miktar/fiyat sıfırken `#DIV/0!` bulunur. Yeni motorda kâr oranı güvenli biçimde 0 döner; parasal parity etkilenmez.",
        "- FORMÜL sayfası aktif modele bağlı değildir.",
        "",
        "## Kullanılan hücrelerin tam dökümü",
        "",
    ]
    for sheet in formula_book.worksheets:
        cached = value_book[sheet.title]
        lines += [f"### {sheet.title}", "", "| Hücre | Tür | Değer / formül | Cached değer |", "|---|---|---|---:|"]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                kind = "Formül" if cell.data_type == "f" else "Sabit"
                lines.append(
                    f"| {cell.coordinate} | {kind} | {markdown_value(cell.value)} | {markdown_value(cached[cell.coordinate].value) if cell.data_type == 'f' else '—'} |"
                )
        lines.append("")

    (DOCS_DIR / "ExcelAnalysis.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_formula_catalog(entries: list[dict[str, Any]]) -> None:
    lines = [
        "# Formula Catalog",
        "",
        "> Kaynak: Legacy Excel formülleri ve workbook içindeki cached sonuçlar. Runtime'da bu formüller çalıştırılmaz.",
        "",
        f"Toplam formül hücresi: **{len(entries)}**.",
        "",
    ]
    for entry in entries:
        classification = "Legacy / Unused" if entry["sheet"] == "FORMÜL" else "Active"
        lines += [
            f"## {entry['sheet']}!{entry['cell']}",
            "",
            f"- Excel Cell: `{entry['sheet']}!{entry['cell']}`",
            f"- Material: {entry['material'] or '—'}",
            f"- Category: {entry['category'] or classification}",
            f"- Excel Formula: `{str(entry['formula']).replace('`', 'ˋ')}`",
            f"- Input Dependencies: {', '.join(f'`{item}`' for item in entry['dependencies']) or '—'}",
            f"- Expected Result: `{markdown_value(entry['expected'])}`",
            f"- C# Rule: `{entry['csharpRule']}`",
            "",
        ]
    (DOCS_DIR / "FormulaCatalog.md").write_text("\n".join(lines), encoding="utf-8")


def write_material_catalog(categories: list[CategoryRange], materials: list[dict[str, Any]]) -> None:
    lines = [
        "# Material Catalog",
        "",
        f"Legacy TEKLİF sayfasından çıkarılan toplam malzeme: **{len(materials)}**.",
        "",
        "| Kod | Grup | Excel satırı | Malzeme | Ölçü | Birim | Miktar kuralı | Fiyat kuralı | Cached miktar | Cached satış toplamı |",
        "|---|---|---:|---|---|---|---|---|---:|---:|",
    ]
    for item in materials:
        lines.append(
            f"| {item['code']} | {item['category']} | {item['excelRow']} | {markdown_value(item['name'])} | {markdown_value(item['specification'])} | {item['unit']} | {item['quantityRuleId']} | {item['pricingRuleId']} | {markdown_value(item['expectedQuantity'])} | {markdown_value(item['expectedSalesLineTotal'])} |"
        )
    lines += [
        "",
        "## Kategori doğrulaması",
        "",
        "| Kod | Ad | Satır sayısı |",
        "|---:|---|---:|",
    ]
    for category in categories:
        count = sum(1 for item in materials if item["categoryId"] == category.code)
        lines.append(f"| {category.code} | {CATEGORY_NAMES[category.code]} | {count} |")
    (DOCS_DIR / "MaterialCatalog.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_lookup_catalog(validations: list[dict[str, str]], formula_book: Any) -> None:
    lookup = formula_book.worksheets[2]
    sections = {
        "Pencere renkleri": [lookup.cell(row, 1).value for row in range(10, 13)],
        "Çatı tipleri": [lookup.cell(row, 12).value for row in range(32, 36)],
        "Kaplama tipleri": [lookup.cell(row, 1).value for row in range(34, 45)],
        "Çatı sistemleri": [lookup.cell(row, 1).value for row in range(45, 47)],
        "Duvar kalınlıkları": [lookup.cell(row, 7).value for row in range(45, 50)],
        "Kat adetleri": [lookup.cell(row, 9).value for row in range(28, 30)],
        "Sorumluluk": [lookup.cell(row, 15).value for row in range(40, 42)],
    }
    lines = ["# Lookup Catalog", ""]
    for title, values in sections.items():
        lines += [f"## {title}", ""]
        lines.extend(f"- {markdown_value(value)}" for value in values if value is not None)
        lines.append("")
    lines += ["## Excel doğrulama kaynakları", "", "| Sayfa | Hedef | Kaynak |", "|---|---|---|"]
    for item in validations:
        lines.append(f"| {item['sheet']} | {item['target']} | {markdown_value(item['formula'])} |")
    (DOCS_DIR / "LookupCatalog.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    try:
        extract()
    except Exception as exc:
        print(f"Extraction failed: {exc}", file=sys.stderr)
        raise
